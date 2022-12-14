import csv
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import doctest
import time

start = time.time()


class Salary:
	"""Класс для представления зарплаты.
	Attributes:
		 salary_from (int): Нижняя граница вилки оклада
		 salary_to (int): Верхняя граница вилки оклада
		 salary_currency (str): Валюта оклада

	"""
	currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
	                   "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

	def __init__(self, salary_from, salary_to, salary_currency):
		"""Инициализирует объект Salary, выполняет конвертацию целочисленных полей
		Args:
			salary_from (int / float / str): Нижняя граница вилки оклада
			salary_to (int / float / str): Верхняя граница вилки оклада
			salary_currency (str): Валюта оклада

		# >>> type(Salary(10.1, 20, 'RUR')).__name__
		# 'Salary'
		# >>> Salary(10.1, 20.2, 'RUR').salary_from
		# 10
		# >>> Salary('1000.1111111111', 20.2, 'RUR').salary_from
		# 1000
		# >>> Salary(10.1, 20.2, 'RUR').salary_to
		# 20
		# >>> Salary(10, '20000.33452543245', 'RUR').salary_to
		# 20000
		# >>> Salary(100, 200, 'EUR').salary_currency
		# 'EUR'
		# >>> Salary(100, 200, 'RUR').salary_currency
		# 'RUR'
		# """
		self.salary_from = int(float(salary_from))
		self.salary_to = int(float(salary_to))
		self.salary_currency = salary_currency

	def get_average(self):
		"""Вычисляет среднюю зарплату из вилки и переводит в рубли с помощью словаря currency_to_rub

		Returns:
			float: Средняя зарплата в рублях
		#
		# >>> Salary(10, 20, 'RUR').get_average()
		# 15.0
		# >>> Salary(100, '150', 'RUR').get_average()
		# 125.0
		# >>> Salary('1000.2', '1500.9999999999', 'RUR').get_average()
		# 1250.0
		# >>> Salary(10, 20, 'EUR').get_average()
		# 898.5
		# >>> Salary(10000, '20000.5', 'UZS').get_average()
		# 82.5
		# """
		return 0.5 * (self.salary_from + self.salary_to) * self.currency_to_rub[self.salary_currency]


class Vacancy:
	"""Класс для представления параметров Вакансии
	Attributes:
		name (str): Название вакансии
		salary (Salary): Оклад вакансии
		area_name (str): Город вакансии
		published_at (str): Дата и время публикации вакансии
	"""

	def __init__(self, vacancy):
		"""Инициализирует объект Vacancy
		Args:
			vacancy (dict): Cловарь с данными

		# >>> type(Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:34:36+0300'})).__name__
		# 'Vacancy'
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:34:36+0300'}).name
		# 'Программист'
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:34:36+0300'}).area_name
		# 'Екатеринбург'
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:34:36+0300'}).published_at
		# '2007-12-03T17:34:36+0300'
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:34:36+0300'}).salary.get_average()
		# 15.0
		# """
		self.name = vacancy['name']
		self.salary = Salary(vacancy['salary_from'], vacancy['salary_to'], vacancy['salary_currency'])
		self.area_name = vacancy['area_name']
		self.published_at = vacancy['published_at']

	def get_year(self):
		"""Возвращает год публикации вакансии в целочисленном виде
		Returns:
			int: Год публикации вакансии
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2007-12-03T17:40:09+0300'}).get_year()
		# 2007
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2022-12-03T17:40:09+0300'}).get_year()
		# 2022
		# >>> Vacancy({'name': 'Программист', 'salary_from': '10', 'salary_to': '20', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at': '2014-12-03T17:40:09+0300'}).get_year()
		# 2014
		# """
		return int(self.published_at[:4])


class DataSet:
	"""Класс для предоставления результатов обработки данных о вакансии
	Attributes:
		file_name (str): Название CSV-файла для обработки данных о конкретной вакансии
		vacancy_name (str): Название вакансии
	"""

	def __init__(self, file_name, name):
		"""Инициализирует объект DataSet
		Args:
			file_name (str): Название CSV-файла для обработки данных о конкретной вакансии
			vacancy_name (str): Название вакансии

		# >>> DataSet('vacancies_by_year.csv', 'Программист').file_name
		# 'vacancies_by_year.csv'
		# >>> DataSet('empty.csv', 'Программист').file_name
		# 'empty.csv'
		# >>> DataSet('any.csv', 'Программист').vacancy_name
		# 'Программист'
		# >>> DataSet('any.csv', 'Аналитик').vacancy_name
		# 'Аналитик'
		# """
		self.file_name = file_name
		self.vacancy_name = name

	def parse_csv(self):
		"""Разбивает данные по словарям для дальнейшей обработки
		Returns:
			1) (dict): Динамика уровня зарплат по годам
			2) (dict): Динамика количества вакансий по годам
			3) (dict): Динамика уровня зарплат по годам для выбранной профессии
			4) (dict): Динамика количества вакансий по годам для выбранной профессии
			5) (dict): Уровень зарплат по городам
			6) (dict): Доля вакансий по городам
			7) (int): Общее число вакансий
		"""
		salary, amount, this_vacancy_salary, this_vacancy_amount, salary_city, share_city = {}, {}, {}, {}, {}, {}
		count = 0
		with open(self.file_name, encoding='utf-8-sig') as file:
			rows = csv.reader(file)
			titles = next(rows)
			for row in rows:
				if len(row) == len(titles) and all(row):
					vacancy = Vacancy(dict(zip(titles, row)))
					count += 1
					year, city = vacancy.get_year(), vacancy.area_name
					self.append_to_dictionary(salary, year, [vacancy.salary.get_average()])
					self.append_to_dictionary(amount, year, 1)
					self.append_to_dictionary(salary_city, city, [vacancy.salary.get_average()])
					self.append_to_dictionary(share_city, city, 1)
					if self.vacancy_name in vacancy.name:
						self.append_to_dictionary(this_vacancy_salary, year, [vacancy.salary.get_average()])
						self.append_to_dictionary(this_vacancy_amount, year, 1)
			return salary, amount, this_vacancy_salary, this_vacancy_amount, salary_city, share_city, count

	@staticmethod
	def append_to_dictionary(dct, key, summand):
		"""Добавление в словарь
			dct (dict): Словарь
			key: Ключ
			summand: Добавляемый объект
		"""
		if key not in dct:
			dct[key] = summand
		else:
			dct[key] += summand

	@staticmethod
	def convert_dct(dct):
		"""Конвертирует значения словаря"""
		for key, value in dct.items():
			dct[key] = 0 if len(value) == 0 else int(sum(value) / len(value))
		return dct

	def get_formatted_data(self, salary, amount, this_vacancy_salary, this_vacancy_amount, salary_city, share_city,
	                       count):
		"""Берет данные из parse_csv и приводит к отформатированному виду
		Returns:
			1) (dict): Динамика уровня зарплат по годам
			2) (dict): Динамика количества вакансий по годам
			3) (dict): Динамика уровня зарплат по годам для выбранной профессии
			4) (dict): Динамика количества вакансий по годам для выбранной профессии
			5) (dict): Уровень зарплат по городам (в порядке убывания)
			6) (dict): Доля вакансий по городам (в порядке убывания)
		"""
		for key, value in share_city.items():
			share_city[key] = round(value / count, 4)
		share_city = list(filter(lambda x: x[-1] > 0.01, [(key, value) for key, value in share_city.items()]))
		salary_city = sorted(
			[(key, value) for key, value in self.convert_dct(salary_city).items() if key in dict(share_city)],
			key=lambda x: x[-1], reverse=True)
		return self.convert_dct(salary), amount, self.convert_dct(this_vacancy_salary), this_vacancy_amount, \
		       dict(salary_city[:10]), dict(sorted(share_city, key=lambda x: x[-1], reverse=True)[:10])


class Report:
	"""Класс для представления отчетов по вакансии
	Attributes:
		salary (dict): Динамика уровня зарплат по годам
		amount (dict): Динамика количества вакансий по годам
		this_vacancy_salary (dict): Динамика уровня зарплат по годам для выбранной профессии
		this_vacancy_amount (dict): Динамика количества вакансий по годам для выбранной профессии
		salary_city (dict): Уровень зарплат по городам (в порядке убывания)
		share_city (dict): Доля вакансий по городам (в порядке убывания)
	"""

	def __init__(self, salary, amount, this_vacancy_salary, this_vacancy_amount, salary_city, share_city):
		"""Инициализирует объект Report
		Args:
			salary (dict):
			amount (dict):
			this_vacancy_salary (dict):
			this_vacancy_amount (dict):
			salary_city (dict):
			share_city (dict):

		# >>> Report({'2007': 20000, '2008': 22000}, {}, {}, {}, {}, {}).salary['2007']
		# 20000
		# >>> Report({}, {'2007': 1000, '2008': 1150}, {}, {}, {}, {}).amount['2008']
		# 1150
		# >>> Report({}, {}, {}, {}, {}, {'Москва': 0.447, 'Санкт-Петербург': 0.214}).share_city['Москва']
		# 0.447
		# """
		self.salary = salary
		self.amount = amount
		self.this_vacancy_salary = this_vacancy_salary
		self.this_vacancy_amount = this_vacancy_amount
		self.salary_city = salary_city
		self.share_city = share_city

	def generate_image(self):
		"""Формирует отчет в виде изображения с графиками"""
		fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
		self.vertical_bar(ax1, 'Уровень зарплат по годам', 'Средняя з/п', self.salary, self.this_vacancy_salary)
		self.vertical_bar(ax2, 'Количества вакансий по годам', 'Количество вакансий', self.amount,
		                  self.this_vacancy_amount)

		ax3.set_title('Уровень зарплат по городам', fontsize=8)
		cities = [name.replace('-', '-\n').replace(' ', '\n') for name in reversed(list(self.salary_city.keys()))]
		ax3.barh(cities, list(reversed(list(self.salary_city.values()))), height=0.4, align='center')
		ax3.xaxis.set_tick_params(labelsize=8)
		ax3.yaxis.set_tick_params(labelsize=6)
		ax3.grid(axis='x')

		ax4.set_title('Доля вакансий по городам', fontsize=8)
		summary = sum(self.share_city.values())
		ax4.pie(list(self.share_city.values()) + [1 - summary], labels=list(self.share_city.keys()) + ['Другие'],
		        textprops={'fontsize': 6})

		plt.tight_layout()
		plt.savefig('graph.png')

	def generate_excel(self):
		"""Формирует отчет в виде .xlsx с таблицами"""
		wb = Workbook()
		font_bold = Font(bold=True)
		side = Side(border_style='thin', color='000000')
		border = Border(left=side, right=side, top=side, bottom=side)
		year_statistics = wb.active
		year_statistics.title = 'Статистика по годам'
		year_data = [['Год', 'Средняя зарплата', f'Средняя зарплата - {dataset.vacancy_name}', 'Количество вакансий',
		              f'Количество вакансий - {dataset.vacancy_name}']]
		for year in self.salary:
			year_data.append([year, self.salary[year], self.this_vacancy_salary[year], self.amount[year],
			                  self.this_vacancy_amount[year]])
		self.fill_excel(year_statistics, year_data)
		city_statistics = wb.create_sheet('Статистика по городам')
		city_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
		for (city_salary, salary), (share_city, share) in zip(self.salary_city.items(), self.share_city.items()):
			city_data.append([city_salary, salary, '', share_city, share])
		self.fill_excel(city_statistics, city_data)
		for i in range(len(city_data)):
			city_statistics[f'E{i + 2}'].number_format = '0.00%'
		self.set_border(year_statistics, border, year_data, 'ABCDE')
		self.set_border(city_statistics, border, city_data, 'ABDE')
		self.set_column_width(year_data, year_statistics)
		self.set_column_width(city_data, city_statistics)
		self.set_font(font_bold, year_statistics, city_statistics)
		wb.save('report.xlsx')

	@staticmethod
	def vertical_bar(ax, title, legend, data, this_vacancy_data):
		"""Cтатический метод настройки диаграммы
		Args:
			ax: График
			title (str): Заголовoк графика
			legend (str): Легенда графика
			data (dict): Данные по всем вакансиям
			this_vacancy_data (dict): Данные по выбранной вакансии
		"""
		ax.set_title(title, fontsize=8)
		first_bar = ax.bar(np.array(list(data)) - 0.2, data.values(), width=0.4)
		this_name_bar = ax.bar(np.array(list(this_vacancy_data.keys())) + 0.2, this_vacancy_data.values(), width=0.4)
		ax.legend((first_bar, this_name_bar), (legend, f'{legend} - {dataset.vacancy_name}'), fontsize=8)
		ax.set_xticks(list(data.keys()), list(data.keys()), rotation=90)
		ax.xaxis.set_tick_params(labelsize=8)
		ax.yaxis.set_tick_params(labelsize=8)
		ax.grid(axis='y')

	@staticmethod
	def fill_excel(sheet, data):
		"""Статический метод для заполнения строками таблицы в .xlsx
		Args:
			sheet: Лист в .xlsx
			data (list[list]): Листов листов для заполнения листа .xlsx построчно
		"""
		for el in data:
			sheet.append(el)

	@staticmethod
	def set_border(sheet, border, data, columns):
		"""Статический метод для установки границ ячейки в .xlsx
		Args:
			sheet: Лист в .xlsx
			border (Border): Граница ячейки
			data (list): Данные
			columns (str): Столбцы
		"""
		for i in range(1, len(data) + 1):
			for column in columns:
				sheet[f'{column}{i}'].border = border

	@staticmethod
	def set_font(font, *sheets):
		"""Установка шрифта в .xlsx
		Args:
			font (Font): Шрифт
			*sheets: Листы файла .xlsx
		"""
		for sheet in sheets:
			for column in 'ABCDE':
				sheet[f'{column}{1}'].font = font

	@staticmethod
	def set_column_width(data, sheet):
		"""Выставление ширины столбца в .xlsx
		Args:
			data (list[List]): Данные
			sheet: Лист .xlsx
		"""
		# from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
		column_widths = []
		for row in data:
			for i, cell in enumerate(row):
				cell = str(cell)
				if len(column_widths) > i:
					if len(cell) > column_widths[i]:
						column_widths[i] = len(cell)
				else:
					column_widths += [len(cell)]
		for i, column_width in enumerate(column_widths, 1):
			sheet.column_dimensions[get_column_letter(i)].width = column_width + 2


dataset = DataSet('data/vacancies_by_year.csv', 'Аналитик')
report = Report(*dataset.get_formatted_data(*dataset.parse_csv()))
print(time.time() - start)
# user_choice = input('Вакансии или статистика? ').lower()
user_choice = 'вакансии'
if user_choice == 'вакансии':
	report.generate_image()
elif user_choice == 'статистика':
	report.generate_excel()
else:
	print('Неправильный формат ввода')

